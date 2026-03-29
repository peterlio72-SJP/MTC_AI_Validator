"""
MatCert Simple Agent
====================
Drop a certificate PDF into the inbox/ folder.
Agent reviews it and saves result to Excel.
That's it. No email. No WhatsApp. No complications.

SETUP:
  1. pip install anthropic python-dotenv openpyxl
  2. Create .env file with:
       ANTHROPIC_API_KEY=sk-ant-your-key-here
  3. python agent.py
  4. Drop certificates into inbox/ folder
"""

import os, time, base64, json, shutil, logging
from pathlib import Path
from datetime import datetime

import anthropic
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

# ── Logging ───────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("agent.log", encoding="utf-8"),
    ],
)
log = logging.getLogger()

# ── Folders ───────────────────────────────────────────
BASE       = Path(__file__).parent
INBOX      = BASE / "inbox"       # drop certificates here
DONE       = BASE / "processed"   # moved here after review
STANDARDS  = BASE / "standards"   # put standard PDFs here
REPORTS    = BASE / "reports"     # Excel saved here
EXCEL      = REPORTS / "reviews.xlsx"

for d in [INBOX, DONE, STANDARDS, REPORTS]:
    d.mkdir(exist_ok=True)

CERT_TYPES = {".pdf", ".jpg", ".jpeg", ".png", ".webp"}

# ── AI ────────────────────────────────────────────────
client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY",""))

def b64(data):
    return base64.standard_b64encode(data).decode()

def to_block(data, media_type, title=None):
    if media_type == "application/pdf":
        b = {"type":"document","source":{"type":"base64","media_type":media_type,"data":b64(data)}}
        if title: b["title"] = title
        return b
    return {"type":"image","source":{"type":"base64","media_type":media_type,"data":b64(data)}}

def media_type(path):
    return {".pdf":"application/pdf",".jpg":"image/jpeg",
            ".jpeg":"image/jpeg",".png":"image/png",
            ".webp":"image/webp"}.get(path.suffix.lower(),"application/pdf")

def parse(raw):
    raw = raw.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"): raw = raw[4:]
    return json.loads(raw.strip())

# ── Review ────────────────────────────────────────────
SCHEMA = """{
  "standard": "full standard name e.g. ASTM A106 Grade B",
  "specification": "e.g. ASTM A106",
  "grade": "e.g. Grade B",
  "source": "Library Document or AI Knowledge",
  "confidence": "HIGH or MEDIUM or LOW",
  "heat_number": "value or null",
  "manufacturer": "value or null",
  "test_date": "value or null",
  "chemical": {
    "ELEMENT": {"found": null, "min": null, "max": null, "status": "PASS or FAIL or MISSING"}
  },
  "mechanical": {
    "PROPERTY": {"found": null, "min": null, "max": null, "unit": "", "status": "PASS or FAIL or MISSING"}
  },
  "impact": {
    "temperature": null, "average_J": null, "required_J": null,
    "status": "PASS or FAIL or MISSING or NA"
  },
  "verdict": "PASS or FAIL or CONDITIONAL",
  "failed": ["list of failed items"],
  "missing": ["list of missing items"],
  "summary": "one paragraph summary"
}"""

def review(filepath):
    data       = filepath.read_bytes()
    mt         = media_type(filepath)
    content    = []
    std_used   = []

    # Load standards from folder
    for std in sorted(STANDARDS.glob("*.pdf")):
        content.append(to_block(std.read_bytes(), "application/pdf",
                                title=f"STANDARD: {std.stem}"))
        std_used.append(std.name)

    # Add the certificate
    content.append(to_block(data, mt))

    if std_used:
        instruction = (
            f"Standard documents provided: {', '.join(std_used)}\n\n"
            "Read these standards to get exact requirements. "
            "Then read the mill certificate and compare every value. "
        )
    else:
        instruction = (
            "No standard documents provided. "
            "Use your expert knowledge of ASME, ASTM, API, EN, GB/T standards. "
        )

    instruction += f"\nRespond ONLY in this JSON (no markdown):\n{SCHEMA}"
    content.append({"type":"text","text":instruction})

    resp = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4096,
        system="You are a senior materials engineer. Review mill certificates against standards precisely.",
        messages=[{"role":"user","content":content}],
    )
    result = parse(resp.content[0].text)
    result["_std_used"] = std_used
    return result

# ── Excel ─────────────────────────────────────────────
COLS = ["Time","File","Standard","Grade","Heat No.",
        "Manufacturer","Test Date","Verdict","Source",
        "Confidence","Failed Items","Missing Items","Summary"]
WIDTHS = [16,25,20,10,14,18,12,12,16,10,40,30,50]

def save_excel(result, filename):
    if EXCEL.exists():
        wb = load_workbook(EXCEL)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reviews"
        hf = PatternFill("solid", fgColor="0A0C10")
        hfont = Font(bold=True, color="00D4AA", size=10)
        for i,(h,w) in enumerate(zip(COLS,WIDTHS),1):
            c = ws.cell(row=1,column=i,value=h)
            c.fill=hf; c.font=hfont
            c.alignment=Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes="A2"

    v      = result.get("verdict","CONDITIONAL")
    colors = {"PASS":"0D2E1F","FAIL":"2E0D12","CONDITIONAL":"2E200D"}
    fcolors= {"PASS":"00C77A","FAIL":"FF4455","CONDITIONAL":"FFB020"}
    rf     = PatternFill("solid", fgColor=colors.get(v,"1A1C22"))
    rfont  = Font(color=fcolors.get(v,"E8EAF0"), size=9)

    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        filename,
        result.get("standard","—"),
        result.get("grade","—"),
        result.get("heat_number") or "—",
        result.get("manufacturer") or "—",
        result.get("test_date") or "—",
        v,
        result.get("source","—"),
        result.get("confidence","—"),
        " | ".join(result.get("failed",[])) or "None",
        " | ".join(result.get("missing",[])) or "None",
        result.get("summary","")[:300],
    ]

    rn = ws.max_row + 1
    for i,val in enumerate(row,1):
        c = ws.cell(row=rn,column=i,value=val)
        c.fill=rf; c.font=rfont
        c.alignment=Alignment(vertical="center",wrap_text=(i>=11))
    ws.row_dimensions[rn].height=18
    wb.save(EXCEL)

# ── Main loop ─────────────────────────────────────────
def run():
    log.info("="*45)
    log.info("  MatCert Simple Agent")
    log.info("="*45)
    log.info(f"📂 Drop certificates into:  {INBOX}")
    log.info(f"📚 Standards folder:        {STANDARDS}")
    log.info(f"📊 Results saved to:        {EXCEL}")
    log.info("-"*45)

    std_count = len(list(STANDARDS.glob("*.pdf")))
    if std_count:
        log.info(f"📚 {std_count} standard(s) loaded from library")
    else:
        log.info("📚 No standards in folder — AI will use built-in knowledge")

    log.info("✅ Watching inbox... Press Ctrl+C to stop")
    log.info("-"*45)

    seen = set()

    while True:
        try:
            for fp in sorted(INBOX.iterdir()):
                if fp.is_file() and fp.suffix.lower() in CERT_TYPES and fp not in seen:
                    seen.add(fp)
                    log.info(f"\n📄 Found: {fp.name}")

                    try:
                        log.info("🔬 Reviewing... (please wait 20-40 sec)")
                        result  = review(fp)
                        verdict = result.get("verdict","?")
                        emoji   = {"PASS":"✅","FAIL":"❌"}.get(verdict,"⚠️")

                        log.info(f"{emoji} Verdict: {verdict}")
                        log.info(f"   Standard:  {result.get('standard','—')}")
                        log.info(f"   Heat No.:  {result.get('heat_number','—')}")

                        if result.get("failed"):
                            log.info(f"   ❌ Failed: {', '.join(result['failed'][:3])}")
                        if result.get("missing"):
                            log.info(f"   ⚠️  Missing: {', '.join(result['missing'][:3])}")

                        save_excel(result, fp.name)
                        log.info(f"📊 Saved to Excel: {EXCEL.name}")

                        # Move to processed
                        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
                        dest = DONE / f"{ts}_{fp.name}"
                        shutil.move(str(fp), str(dest))
                        log.info(f"📁 Moved to processed/")
                        log.info("-"*45)

                    except Exception as e:
                        log.error(f"❌ Error: {e}")
                        shutil.move(str(fp), str(DONE / fp.name))

            time.sleep(5)

        except KeyboardInterrupt:
            log.info("\n🛑 Agent stopped")
            break

if __name__ == "__main__":
    if not os.getenv("ANTHROPIC_API_KEY"):
        log.error("❌ ANTHROPIC_API_KEY not set in .env file")
    else:
        run()
